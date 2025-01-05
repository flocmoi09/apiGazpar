import logging
from datetime import datetime, time,timedelta
import pytz
import dateparser
from pygazpar.enum import Frequency,PropertyName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from typing import  List, Dict
from pygazpar.enum import NatureReleve, QualificationReleve, StatusReleve
from pygazpar.types.releves_result_type import Releves_result_type
from pygazpar.types.consommation_type import Releves_type
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta

FIRST_DATA_LINE_NUMBER = 10

Logger = logging.getLogger(__name__)


# ------------------------------------------------------------------------------------------------------------
class ExcelParser:
    OUTPUT_DATE_FORMAT = "%Y-%m-%d"
    OUTPUT_DATETIME_FORMAT = "%Y-%m-%dT%H:%M:%S%z"

    INPUT_DATE_FORMAT = "%d/%m/%Y"
    # ------------------------------------------------------
    @staticmethod
    def parse(dataFilename: str, dataReadingFrequency: Frequency) -> List[Releves_result_type]:

        parseByFrequency = {
            Frequency.HOURLY: ExcelParser.__parseHourly,
            Frequency.DAILY: ExcelParser.__parseDaily,
            Frequency.WEEKLY: ExcelParser.__parseWeekly,
            Frequency.MONTHLY: ExcelParser.__parseMonthly
        }

        Logger.debug(f"Loading Excel data file '{dataFilename}'...")

        workbook = load_workbook(filename=dataFilename)

        worksheet = workbook.active

        res = parseByFrequency[dataReadingFrequency](worksheet)  # type: ignore

        workbook.close()

        return res

    # ------------------------------------------------------
    @staticmethod
    def __fillRow(row: Dict, propertyName: str, cell: Cell, isNumber: bool):

        if cell.value is not None:
            if isNumber:
                if type(cell.value) is str:
                    if len(cell.value.strip()) > 0:
                        row[propertyName] = float(cell.value.replace(',', '.'))
                else:
                    row[propertyName] = cell.value
            else:
                row[propertyName] = cell.value.strip() if type(cell.value) is str else cell.value
        else:
            row[propertyName] = None

    # ------------------------------------------------------
    @staticmethod
    def __parseHourly(worksheet: Worksheet) -> List[Releves_result_type]:
        return []

    # ------------------------------------------------------
    @staticmethod
    def __parseDaily(worksheet: Worksheet) -> List[Releves_result_type]:

        res = []
       
        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()

        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                date_journee = datetime.strptime(worksheet.cell(column=2, row=rownum).value, ExcelParser.INPUT_DATE_FORMAT).date()
                info=pytz.timezone('Europe/Paris')
                MyTime = time(6, 0, 0)  #hr/min/sec
                datetime_debut = datetime.combine(date_journee, MyTime)
                datetime_debut_localize=info.localize(datetime_debut)
                row[PropertyName.JOURNEE_GAZIERE.value] = date_journee.strftime(ExcelParser.OUTPUT_DATE_FORMAT)
                row[PropertyName.DATE_DEBUT.value]= datetime_debut_localize.isoformat()
                row[PropertyName.DATE_FIN.value]= (datetime_debut_localize+timedelta(days=1)).isoformat()

                ExcelParser.__fillRow(row, PropertyName.START_INDEX.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.END_INDEX.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=6, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.CONVERTER_FACTOR.value, worksheet.cell(column=7, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=8, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.QUALIFICATION.value, worksheet.cell(column=9, row=rownum), False)  # type: ignore
                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                releve = Releves_type(**row)
                releve_result = Releves_result_type(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)

                res.append(releve_result)

        Logger.debug(f"Daily data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parseWeekly(worksheet: Worksheet) -> List[Releves_result_type]:

        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()
        info=pytz.timezone('Europe/Paris')
        MyTime = time(6, 0, 0)  #hr/min/sec
        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                dateField=worksheet.cell(column=2, row=rownum).value
                dateStart=dateField.split('au')[0]
                dateEnd=dateField.split('au')[1]
                dateStartDT=parse(dateStart, fuzzy_with_tokens=True)
                dateEndDT=parse(dateEnd, fuzzy_with_tokens=True)
                dateStartDT = datetime.combine(dateStartDT[0], MyTime)
                dateEndDT = datetime.combine(dateEndDT[0], MyTime)
                dateStartDT=info.localize(dateStartDT)
                dateEndDT=info.localize(dateEndDT)

                row[PropertyName.DATE_DEBUT.value]= dateStartDT.isoformat()
                row[PropertyName.DATE_FIN.value]= (dateEndDT+timedelta(days=1)).isoformat()
                row[PropertyName.JOURNEE_GAZIERE.value] =None
                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore

                row[PropertyName.START_INDEX.value]=None
                row[PropertyName.END_INDEX.value]=None
                row[PropertyName.CONVERTER_FACTOR.value]=None
                row[PropertyName.QUALIFICATION.value]=QualificationReleve.ESTIME.value

                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                releve = Releves_type(**row)
                releve_result = Releves_result_type(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)
                res.append(releve_result)

        Logger.debug(f"Weekly data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parseMonthly(worksheet: Worksheet) -> List[Releves_result_type]:

        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()
        info=pytz.timezone('Europe/Paris')
       
        MyTime = time(6, 0, 0)  #hr/min/sec
        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                dateField=worksheet.cell(column=2, row=rownum).value
                dateStartDT=dateparser.parse(dateField, locales=['fr'])
                dateStartDT=dateStartDT.replace(day=1)
                dateStartDT = datetime.combine(dateStartDT, MyTime)
                dateStartDT=info.localize(dateStartDT)
                row[PropertyName.DATE_DEBUT.value]= dateStartDT.isoformat()
                row[PropertyName.DATE_FIN.value]= (dateStartDT+relativedelta(months=1)).isoformat()
                row[PropertyName.JOURNEE_GAZIERE.value] =None

                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore

                row[PropertyName.START_INDEX.value]=None
                row[PropertyName.END_INDEX.value]=None
                row[PropertyName.CONVERTER_FACTOR.value]=None
                row[PropertyName.QUALIFICATION.value]=QualificationReleve.ESTIME.value
                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                
                releve = Releves_type(**row)
                releve_result = Releves_result_type(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)
                res.append(releve_result)

        Logger.debug(f"Monthly data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res
